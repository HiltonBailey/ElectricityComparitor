#!/usr/bin/env python3
"""
Phase 1: Extend Retailer_Rates.
Saves to temp file to avoid corruption on timeout.
"""
import openpyxl
import os
import shutil

EXCEL_FILE = "HA Energy 5 Min Pricing.xlsx"
TEMP_FILE = "HA Energy 5 Min Pricing temp.xlsx"

# Work on temp copy
shutil.copy2(EXCEL_FILE, TEMP_FILE)
wb = openpyxl.load_workbook(TEMP_FILE)
ws = wb['Retailer_Rates']

# Add new headers
ws['Q1'] = 'Pricing_Model'
ws['R1'] = 'Subscription_Monthly'
ws['S1'] = 'Subscription_Daily'

# Update existing rows with Pricing_Model
for row, model in [(2, 'fixed_tou'), (3, 'fixed_tou'), (4, 'fixed_tou')]:
    ws[f'Q{row}'] = model
    ws[f'R{row}'] = 0
    ws[f'S{row}'] = 0

# Add FlowPower entry (row 5)
r = 5
ws[f'A{r}'] = 1
ws[f'B{r}'] = 'FlowPower'
ws[f'C{r}'] = 1.3419
ws[f'D{r}'] = 0.34
ws[f'E{r}'] = 0
ws[f'F{r}'] = 0
ws[f'G{r}'] = 0.45
ws[f'H{r}'] = 0
ws[f'I{r}'] = 0
ws[f'J{r}'] = 17.5
ws[f'K{r}'] = 19.5
ws[f'L{r}'] = 0
ws[f'M{r}'] = 0
ws[f'N{r}'] = 0
ws[f'O{r}'] = 0
ws[f'P{r}'] = 0
ws[f'Q{r}'] = 'hybrid'
ws[f'R{r}'] = 0
ws[f'S{r}'] = 0

# Add Amber entry (row 6)
r = 6
ws[f'A{r}'] = 5
ws[f'B{r}'] = 'Amber'
ws[f'C{r}'] = 1.76
ws[f'D{r}'] = 0
ws[f'E{r}'] = 0
ws[f'F{r}'] = 0
ws[f'G{r}'] = 0
ws[f'H{r}'] = 0
ws[f'I{r}'] = 0
ws[f'J{r}'] = 0
ws[f'K{r}'] = 0
ws[f'L{r}'] = 0
ws[f'M{r}'] = 0
ws[f'N{r}'] = 0
ws[f'O{r}'] = 0
ws[f'P{r}'] = 0
ws[f'Q{r}'] = 'variable'
ws[f'R{r}'] = 25
ws[f'S{r}'] = round(25 / 30, 4)

wb.save(TEMP_FILE)
wb.close()

# Replace original
os.replace(TEMP_FILE, EXCEL_FILE)
print("Phase 1 complete: Retailer_Rates extended")
