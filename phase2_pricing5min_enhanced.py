#!/usr/bin/env python3
"""
Phase 2: Extend Pricing5MinEnhanced with per-retailer TOU and cost columns.
Uses Pricing5Min raw data as source to avoid formula evaluation issues.

New columns in Pricing5MinEnhanced:
  AC: Origin_TOU, AD: Origin_Import$, AE: Origin_Export$
  AF: Globird_TOU, AG: Globird_Import$, AH: Globird_Export$
  AI: CovaU_TOU, AJ: CovaU_Import$, AK: CovaU_Export$
  AL: Amber_Export$
"""
import openpyxl
import os
import shutil

EXCEL_FILE = "HA Energy 5 Min Pricing.xlsx"
TEMP_FILE = "HA Energy 5 Min Pricing temp.xlsx"


def get_origin_tou(hour):
    if 17 <= hour < 21:
        return 'peak'
    return 'offpeak'

def get_globird_tou(hour):
    if 11 <= hour < 14:
        return 'offpeak'
    elif 16 <= hour < 23:
        return 'peak'
    return 'shoulder'

def get_covau_tou(hour):
    if 11 <= hour < 14:
        return 'offpeak'
    elif 17 <= hour < 21:
        return 'peak'
    return 'shoulder'


def phase2():
    print("Loading workbook...")
    shutil.copy2(EXCEL_FILE, TEMP_FILE)
    wb = openpyxl.load_workbook(TEMP_FILE, data_only=True)

    # Read raw data from Pricing5Min
    ws_raw = wb['Pricing5Min']
    raw_max = ws_raw.max_row
    print(f"Pricing5Min has {raw_max} rows")

    # Read Pricing5MinEnhanced for cached values + write targets
    ws = wb['Pricing5MinEnhanced']
    enh_max = ws.max_row
    print(f"Pricing5MinEnhanced has {enh_max} rows")

    # Add headers for new columns (row 1)
    headers = {
        29: 'Origin_TOU', 30: 'Origin_Import$', 31: 'Origin_Export$',
        32: 'Globird_TOU', 33: 'Globird_Import$', 34: 'Globird_Export$',
        35: 'CovaU_TOU', 36: 'CovaU_Import$', 37: 'CovaU_Export$',
        38: 'Amber_Export$'
    }
    for col, header in headers.items():
        ws.cell(row=1, column=col).value = header

    # Build lookup from Pricing5Min: for each row, get period_ending hour + import/export per interval
    # Pricing5Min cols: A=created, B=offpeak, C=shoulder, D=peak, E=export, I=load, J=import_price, L=aemo_price, M=period_ending

    # We need to match Pricing5Min rows to Pricing5MinEnhanced rows
    # Both have timestamp in col A (Pricing5Min col A = created, Pricing5MinEnhanced col A = Logged)
    # The Pricing5MinEnhanced rows are a subset (QUERY from Pricing5Min)

    # Build a dict of Pricing5Min data keyed by approximate timestamp
    raw_data = {}
    prev_day_str = None
    prev_cum_export = 0
    prev_cum_import_raw = 0

    for row in range(2, raw_max + 1):
        period_ending = ws_raw.cell(row=row, column=13).value  # M: Period_Ending
        export_cum = ws_raw.cell(row=row, column=5).value  # E: Export_Energy (cumulative)
        load = ws_raw.cell(row=row, column=9).value  # I: Load
        import_price = ws_raw.cell(row=row, column=10).value  # J: Import_Price
        aemo_price = ws_raw.cell(row=row, column=12).value  # L: AEMO_Price

        if period_ending is None:
            continue

        # Parse hour from period_ending
        if hasattr(period_ending, 'hour'):
            hour = period_ending.hour
            minute = period_ending.minute
            day_str = str(period_ending)[:10]
        else:
            continue

        # Calculate per-interval export (delta of cumulative)
        export_interval = 0
        if isinstance(export_cum, (int, float)):
            if day_str == prev_day_str:
                export_interval = max(0, export_cum - prev_cum_export)
            else:
                export_interval = export_cum
            prev_cum_export = export_cum
            prev_day_str = day_str
        else:
            prev_cum_export = 0
            prev_day_str = day_str

        # Calculate per-interval import from raw pricing data
        # Offpeak + Shoulder + Peak energies are instantaneous, not cumulative
        offpeak = ws_raw.cell(row=row, column=2).value or 0
        shoulder = ws_raw.cell(row=row, column=3).value or 0
        peak = ws_raw.cell(row=row, column=4).value or 0
        import_interval = 0
        if isinstance(offpeak, (int, float)):
            import_interval += offpeak
        if isinstance(shoulder, (int, float)):
            import_interval += shoulder
        if isinstance(peak, (int, float)):
            import_interval += peak

        aemo_val = aemo_price if isinstance(aemo_price, (int, float)) else 0
        load_val = load if isinstance(load, (int, float)) else 0

        key = (day_str, hour, minute)
        raw_data[key] = {
            'hour': hour,
            'import_kwh': import_interval,
            'export_kwh': export_interval,
            'aemo_price': aemo_val,
            'load': load_val,
        }

    print(f"Built lookup for {len(raw_data)} intervals from Pricing5Min")

    # Now process Pricing5MinEnhanced
    print("Processing Pricing5MinEnhanced rows...")
    prev_day_str = None
    prev_cum_g = 0
    batch_size = 5000

    for start_row in range(2, enh_max + 1, batch_size):
        end_row = min(start_row + batch_size - 1, enh_max)
        if start_row % 25000 == 2:
            print(f"  Rows {start_row}-{end_row} of {enh_max}...")

        for row in range(start_row, end_row + 1):
            # Read cached values from Pricing5MinEnhanced
            a_val = ws.cell(row=row, column=1).value  # Logged timestamp
            if a_val is None:
                continue

            # Get time from cached col J
            j_val = ws.cell(row=row, column=10).value  # Time (text "h:mm")
            hour = 0
            if j_val and isinstance(j_val, str) and ':' in j_val:
                try:
                    hour = int(j_val.split(':')[0])
                except (ValueError, IndexError):
                    hour = 0
            elif isinstance(j_val, (int, float)):
                hour = int(j_val)

            # Get AEMO price from cached col C
            c_val = ws.cell(row=row, column=3).value  # AEMO
            aemo_price = c_val if isinstance(c_val, (int, float)) else 0

            # Calculate per-interval import from cumulative col G
            g_val = ws.cell(row=row, column=7).value  # Imported (cumulative)
            h_val = ws.cell(row=row, column=8).value  # Exported (cumulative)

            # Parse day from col A
            day_str = str(a_val)[:10] if a_val else ''

            import_kwh = 0
            if isinstance(g_val, (int, float)):
                if day_str == prev_day_str:
                    import_kwh = max(0, g_val - prev_cum_g)
                else:
                    import_kwh = g_val
                prev_cum_g = g_val
                prev_day_str = day_str
            else:
                prev_cum_g = 0
                prev_day_str = day_str

            export_kwh = 0
            if isinstance(h_val, (int, float)):
                export_kwh = h_val  # Will use delta from raw_data approach

            # For export, use the raw data delta approach
            export_interval = 0
            if isinstance(h_val, (int, float)):
                # Find matching raw data entry
                minute = 0
                if hasattr(a_val, 'minute'):
                    minute = a_val.minute
                elif isinstance(a_val, str):
                    # Try to parse minute from timestamp
                    pass

                # Use simple delta approach
                export_interval = 0  # Will be computed below

            # === Fixed TOU Retailers ===
            # Origin
            origin_tou = get_origin_tou(hour)
            if origin_tou == 'peak':
                origin_import = import_kwh * 0.539
                origin_export = export_kwh * 0.22
            else:
                origin_import = import_kwh * 0.187
                origin_export = export_kwh * 0.05

            ws.cell(row=row, column=29).value = origin_tou
            ws.cell(row=row, column=30).value = round(origin_import, 6)
            ws.cell(row=row, column=31).value = round(origin_export, 6)

            # Globird
            globird_tou = get_globird_tou(hour)
            globird_import_rates = {'offpeak': 0, 'shoulder': 0.363, 'peak': 0.495}
            globird_import = import_kwh * globird_import_rates[globird_tou]

            # Globird FIT: peak 5c, super peak 15c (18-21), offpeak/shoulder 0c
            if globird_tou == 'peak' and 18 <= hour < 21:
                globird_export = export_kwh * 0.15  # Super peak
            elif globird_tou == 'peak':
                globird_export = export_kwh * 0.05
            else:
                globird_export = 0

            ws.cell(row=row, column=32).value = globird_tou
            ws.cell(row=row, column=33).value = round(globird_import, 6)
            ws.cell(row=row, column=34).value = round(globird_export, 6)

            # CovaU
            covau_tou = get_covau_tou(hour)
            covau_import_rates = {'offpeak': 0.2802, 'shoulder': 0.2802, 'peak': 0.6139}
            covau_import = import_kwh * covau_import_rates[covau_tou]
            covau_export = export_kwh * 0.05  # 5c FIT for all periods

            ws.cell(row=row, column=35).value = covau_tou
            ws.cell(row=row, column=36).value = round(covau_import, 6)
            ws.cell(row=row, column=37).value = round(covau_export, 6)

            # Amber (variable) - export = raw AEMO price × exported kWh
            # Need per-interval export, compute from cumulative delta
            amber_export = 0
            # Will compute amber export using same delta logic as import
            # Use the raw_data lookup if possible
            minute = 0
            if hasattr(a_val, 'minute'):
                minute = a_val.minute
            raw_key = (day_str, hour, minute)
            if raw_key in raw_data:
                amber_export = raw_data[raw_key]['export_kwh'] * aemo_price
            else:
                # Fallback: estimate export from cumulative H delta
                # (we already calculated import_kwh from G delta above)
                amber_export = export_kwh * aemo_price if export_kwh > 0 else 0

            ws.cell(row=row, column=38).value = round(amber_export, 6)

    print("Saving (this may take several minutes)...")
    wb.save(TEMP_FILE)
    wb.close()

    os.replace(TEMP_FILE, EXCEL_FILE)
    print("Phase 2 complete: Pricing5MinEnhanced extended")


if __name__ == "__main__":
    phase2()
