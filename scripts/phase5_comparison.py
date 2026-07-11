#!/usr/bin/env python3
"""
Phase 5: Create Comparison summary tab with billing period,
rolling 30-day, and smart export comparison.
"""
import openpyxl
import os
import shutil
from datetime import datetime, date, timedelta

EXCEL_FILE = "HA Energy 5 Min Pricing.xlsx"
TEMP_FILE = "HA Energy 5 Min Pricing temp.xlsx"


def phase5():
    print("Loading workbook...")
    shutil.copy2(EXCEL_FILE, TEMP_FILE)
    wb = openpyxl.load_workbook(TEMP_FILE, data_only=True)

    # Read Retailer_Comparison data
    ws_rc = wb['Retailer_Comparison']
    rc_max = ws_rc.max_row
    print(f"Retailer_Comparison: {rc_max} rows")

    # Read SmartExport data
    ws_se = wb['SmartExport']
    se_max = ws_se.max_row
    print(f"SmartExport: {se_max} rows")

    # Create Comparison tab
    if 'Comparison' in wb.sheetnames:
        del wb['Comparison']

    ws = wb.create_sheet('Comparison')

    # ===== SECTION 1: CURRENT BILLING PERIOD =====
    ws['A1'] = 'CURRENT BILLING PERIOD COMPARISON'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

    # Billing period dates (from PricingDaily - last 30 days)
    ws['A3'] = 'Period:'
    ws['B3'] = 'Last 30 days (from most recent)'

    # Headers
    headers = ['Metric', 'Flow', 'Origin', 'Globird', 'CovaU', 'Amber']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = h
        cell.font = openpyxl.styles.Font(bold=True)

    # Collect last 30 days data
    retailers = ['Flow', 'Origin', 'Globird', 'CovaU', 'Amber']
    metric_cols = {
        'Flow': (2, 'E'),    # Net$
        'Origin': (9, 'I'),
        'Globird': (13, 'M'),
        'CovaU': (17, 'Q'),
        'Amber': (22, 'V'),
    }

    totals = {r: {'import': 0, 'export': 0, 'net': 0, 'days': 0} for r in retailers}

    for row in range(2, min(rc_max + 1, 32)):  # Last 30 days
        for r in retailers:
            col_map = {'Flow': 2, 'Origin': 6, 'Globird': 10, 'CovaU': 14, 'Amber': 18}
            net_col = {'Flow': 5, 'Origin': 9, 'Globird': 13, 'CovaU': 17, 'Amber': 22}
            imp_col = {'Flow': 2, 'Origin': 6, 'Globird': 10, 'CovaU': 14, 'Amber': 18}
            exp_col = {'Flow': 3, 'Origin': 7, 'Globird': 11, 'CovaU': 15, 'Amber': 19}

            imp = ws_rc.cell(row=row, column=imp_col[r]).value
            exp = ws_rc.cell(row=row, column=exp_col[r]).value
            net = ws_rc.cell(row=row, column=net_col[r]).value

            if isinstance(imp, (int, float)):
                totals[r]['import'] += imp
            if isinstance(exp, (int, float)):
                totals[r]['export'] += exp
            if isinstance(net, (int, float)):
                totals[r]['net'] += net
                totals[r]['days'] += 1

    # Write metrics
    metrics = ['Total Import $', 'Total Export $', 'Net Cost $', 'Avg Daily Net $', 'Days']
    for i, metric in enumerate(metrics):
        r = 6 + i
        ws.cell(row=r, column=1).value = metric
        for j, retailer in enumerate(retailers):
            col = j + 2
            if metric == 'Total Import $':
                ws.cell(row=r, column=col).value = round(totals[retailer]['import'], 2)
            elif metric == 'Total Export $':
                ws.cell(row=r, column=col).value = round(totals[retailer]['export'], 2)
            elif metric == 'Net Cost $':
                ws.cell(row=r, column=col).value = round(totals[retailer]['net'], 2)
            elif metric == 'Avg Daily Net $':
                days = max(totals[retailer]['days'], 1)
                ws.cell(row=r, column=col).value = round(totals[retailer]['net'] / days, 2)
            elif metric == 'Days':
                ws.cell(row=r, column=col).value = totals[retailer]['days']

    # Savings vs Flow
    r = 12
    ws.cell(row=r, column=1).value = 'Savings vs Flow $'
    ws.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
    flow_net = totals['Flow']['net']
    for j, retailer in enumerate(retailers):
        col = j + 2
        savings = flow_net - totals[retailer]['net']
        ws.cell(row=r, column=col).value = round(savings, 2)
        if savings > 0:
            ws.cell(row=r, column=col).font = openpyxl.styles.Font(color='008000')
        elif savings < 0:
            ws.cell(row=r, column=col).font = openpyxl.styles.Font(color='FF0000')

    # ===== SECTION 2: SMART EXPORT COMPARISON =====
    r = 15
    ws.cell(row=r, column=1).value = 'SMART EXPORT OPTIMISATION COMPARISON'
    ws.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    r = 17
    ws.cell(row=r, column=1).value = 'Optimised strategy assumes:'
    ws.cell(row=r + 1, column=1).value = '  - Export to max capacity during highest FIT value hours'
    ws.cell(row=r + 2, column=1).value = '  - Retain 15% of avg daily solar for next-day consumption'
    ws.cell(row=r + 3, column=1).value = '  - Flow: export during 17:30-19:30 window only'

    r = 22
    smart_headers = ['Metric', 'Flow', 'Origin', 'Globird', 'CovaU', 'Amber']
    for col, h in enumerate(smart_headers, 1):
        cell = ws.cell(row=r, column=col)
        cell.value = h
        cell.font = openpyxl.styles.Font(bold=True)

    # Collect smart export totals
    smart_totals = {ret: {'export_kwh': 0, 'export_value': 0, 'import_kwh': 0, 'import_value': 0, 'net': 0, 'days': 0}
                    for ret in retailers}

    col_offsets = {
        'Flow': {'exp_kwh': 5, 'exp_val': 6, 'imp_kwh': 7, 'imp_val': 8, 'net': 9},
        'Origin': {'exp_kwh': 10, 'exp_val': 11, 'imp_kwh': 12, 'imp_val': 13, 'net': 14},
        'Globird': {'exp_kwh': 15, 'exp_val': 16, 'imp_kwh': 17, 'imp_val': 18, 'net': 19},
        'CovaU': {'exp_kwh': 20, 'exp_val': 21, 'imp_kwh': 22, 'imp_val': 23, 'net': 24},
        'Amber': {'exp_kwh': 25, 'exp_val': 26, 'imp_kwh': 27, 'imp_val': 28, 'net': 29},
    }

    key_map = {'export_kwh': 'exp_kwh', 'export_value': 'exp_val',
               'import_kwh': 'imp_kwh', 'import_value': 'imp_val', 'net': 'net'}

    for row in range(2, se_max + 1):
        for ret in retailers:
            cols = col_offsets[ret]
            for tot_key, col_key in key_map.items():
                val = ws_se.cell(row=row, column=cols[col_key]).value
                if isinstance(val, (int, float)):
                    smart_totals[ret][tot_key] += val
            smart_totals[ret]['days'] += 1

    # Write smart export metrics
    smart_metrics = [
        ('Total Optimised Export kWh', 'export_kwh'),
        ('Total Optimised Export $', 'export_value'),
        ('Total Optimised Import kWh', 'import_kwh'),
        ('Total Optimised Import $', 'import_value'),
        ('Net Cost $', 'net'),
        ('Avg Daily Net $', 'net'),
    ]

    for i, (metric, key) in enumerate(smart_metrics):
        r = 23 + i
        ws.cell(row=r, column=1).value = metric
        for j, retailer in enumerate(retailers):
            col = j + 2
            val = smart_totals[retailer][key]
            if key == 'net' and metric == 'Avg Daily Net $':
                days = max(smart_totals[retailer]['days'], 1)
                val = val / days
            ws.cell(row=r, column=col).value = round(val, 2)

    # Smart export savings vs Flow
    r = 30
    ws.cell(row=r, column=1).value = 'Smart Export Savings vs Flow $'
    ws.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True)
    flow_smart_net = smart_totals['Flow']['net']
    for j, retailer in enumerate(retailers):
        col = j + 2
        savings = flow_smart_net - smart_totals[retailer]['net']
        ws.cell(row=r, column=col).value = round(savings, 2)
        if savings > 0:
            ws.cell(row=r, column=col).font = openpyxl.styles.Font(color='008000')
        elif savings < 0:
            ws.cell(row=r, column=col).font = openpyxl.styles.Font(color='FF0000')

    # ===== SECTION 3: RATE COMPARISON =====
    r = 33
    ws.cell(row=r, column=1).value = 'RETAILER RATE COMPARISON'
    ws.cell(row=r, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    r = 35
    rate_headers = ['Rate Type', 'Flow', 'Origin', 'Globird', 'CovaU', 'Amber']
    for col, h in enumerate(rate_headers, 1):
        cell = ws.cell(row=r, column=col)
        cell.value = h
        cell.font = openpyxl.styles.Font(bold=True)

    rates = [
        ('Daily Supply Charge', 1.3419, 1.2567/365, 1.32/365, 1.30/365, 1.76),
        ('Monthly Subscription', 0, 0, 0, 0, 25),
        ('Offpeak Rate $/kWh', '0.34-PEA', 0.187, 0, 0.2802, 'AEMO+NW'),
        ('Shoulder Rate $/kWh', 'N/A', 0.187, 0.363, 0.2802, 'AEMO+NW'),
        ('Peak Rate $/kWh', '0.34-PEA', 0.539, 0.495, 0.6139, 'AEMO+NW'),
        ('Export Rate $/kWh', 0.45, 0.05, 0.05, 0.05, 'AEMO'),
        ('Export Window', '17:30-19:30', 'All hours', 'All hours', 'All hours', 'All hours'),
        ('Peak Export Rate', 0.45, 0.22, 0.05, 0.05, 'AEMO'),
        ('Super Peak FIT', 'N/A', 'N/A', 0.15, 0.18, 'N/A'),
    ]

    for i, (metric, *vals) in enumerate(rates):
        r = 36 + i
        ws.cell(row=r, column=1).value = metric
        for j, val in enumerate(vals):
            col = j + 2
            if isinstance(val, float):
                ws.cell(row=r, column=col).value = round(val, 4)
            else:
                ws.cell(row=r, column=col).value = val

    # Set column widths
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    print("Saving...")
    wb.save(TEMP_FILE)
    wb.close()

    os.replace(TEMP_FILE, EXCEL_FILE)
    print("Phase 5 complete: Comparison tab created")


if __name__ == "__main__":
    phase5()
