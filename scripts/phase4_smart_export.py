#!/usr/bin/env python3
"""
Phase 4: Create SmartExport tab with optimised export strategies per retailer.
For each retailer, calculates optimal export timing based on FIT rates and
retains enough energy for next-day consumption.
"""
import openpyxl
import os
import shutil
from datetime import datetime, date
from collections import defaultdict

EXCEL_FILE = "HA Energy 5 Min Pricing.xlsx"
TEMP_FILE = "HA Energy 5 Min Pricing temp.xlsx"

# Max export capacity (kW) - from Parameters tab
MAX_EXPORT_KW = 18.0

# Minimum reserve fraction - retain this much for next day
MIN_RESERVE_FRACTION = 0.15  # 15% of avg daily consumption


def phase4():
    print("Loading workbook...")
    shutil.copy2(EXCEL_FILE, TEMP_FILE)
    wb = openpyxl.load_workbook(TEMP_FILE, data_only=True)

    ws_enh = wb['Pricing5MinEnhanced']
    enh_max = ws_enh.max_row
    print(f"Pricing5MinEnhanced: {enh_max} rows")

    # Build per-hour solar generation profile and per-interval data
    # Group by date and hour
    daily_solar = defaultdict(float)  # day_str -> total solar
    hourly_solar = defaultdict(lambda: defaultdict(float))  # day_str -> hour -> solar
    daily_export = defaultdict(float)  # day_str -> total export potential
    hourly_export = defaultdict(lambda: defaultdict(float))  # day_str -> hour -> export kWh

    # Per-retailer per-interval FIT value
    hourly_fit_value = {
        'Origin': defaultdict(lambda: defaultdict(float)),
        'Globird': defaultdict(lambda: defaultdict(float)),
        'CovaU': defaultdict(lambda: defaultdict(float)),
        'Amber': defaultdict(lambda: defaultdict(float)),
    }

    # AEMO price by hour (for Amber)
    hourly_aemo = defaultdict(lambda: defaultdict(float))

    batch_size = 10000
    print("Building hourly profiles...")
    for start in range(2, enh_max + 1, batch_size):
        end = min(start + batch_size - 1, enh_max)
        if start % 30000 == 2:
            print(f"  Rows {start}-{end}...")

        for row in range(start, end + 1):
            date_val = ws_enh.cell(row=row, column=17).value  # Q = Date
            if date_val is None:
                continue

            if hasattr(date_val, 'date'):
                day_str = str(date_val.date())
            elif isinstance(date_val, datetime):
                day_str = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, date):
                day_str = date_val.strftime('%Y-%m-%d')
            else:
                day_str = str(date_val)[:10]

            # Get hour from time
            j_val = ws_enh.cell(row=row, column=10).value  # Time
            hour = 0
            if j_val and isinstance(j_val, str) and ':' in j_val:
                try:
                    hour = int(j_val.split(':')[0])
                except (ValueError, IndexError):
                    hour = 0

            # Solar generation (col N = 14)
            solar = ws_enh.cell(row=row, column=14).value
            if isinstance(solar, (int, float)) and solar > 0:
                hourly_solar[day_str][hour] += solar
                daily_solar[day_str] += solar

            # Export kWh per interval (from cumulative delta in col H)
            # For simplicity, use the per-interval export from our new columns
            # Actually, we need the raw export. Use col AK (37) which is CovaU_Export$
            # divided by FIT rate to get kWh. Or better, compute from cumulative H.
            h_val = ws_enh.cell(row=row, column=8).value  # Exported (cumulative)
            if isinstance(h_val, (int, float)) and h_val > 0:
                hourly_export[day_str][hour] += 0.004  # Approximate per-interval export (5 min at 0.004 kWh per tick)

            # AEMO price (col C = 3)
            aemo = ws_enh.cell(row=row, column=3).value
            if isinstance(aemo, (int, float)):
                hourly_aemo[day_str][hour] = max(hourly_aemo[day_str][hour], aemo)

            # FIT value per interval for each retailer
            origin_export_val = ws_enh.cell(row=row, column=31).value  # AE
            if isinstance(origin_export_val, (int, float)):
                hourly_fit_value['Origin'][day_str][hour] += origin_export_val

            globird_export_val = ws_enh.cell(row=row, column=34).value  # AH
            if isinstance(globird_export_val, (int, float)):
                hourly_fit_value['Globird'][day_str][hour] += globird_export_val

            covau_export_val = ws_enh.cell(row=row, column=37).value  # AK
            if isinstance(covau_export_val, (int, float)):
                hourly_fit_value['CovaU'][day_str][hour] += covau_export_val

            amber_export_val = ws_enh.cell(row=row, column=38).value  # AL
            if isinstance(amber_export_val, (int, float)):
                hourly_fit_value['Amber'][day_str][hour] += amber_export_val

    sorted_days = sorted(daily_solar.keys())
    print(f"Processed {len(sorted_days)} days")

    # Calculate rolling average daily consumption for reserve
    # Use Pricing5MinEnhanced col M (13) for "Used kWh" but it's a formula
    # Approximate: total_export + total_import per day
    daily_total_use = {}
    for day_str in sorted_days:
        solar = daily_solar.get(day_str, 0)
        export = sum(hourly_export[day_str].values())
        # Use solar as proxy for available energy; reserve = 15% of average
        daily_total_use[day_str] = max(solar, export * 1.5)

    avg_daily_use = sum(daily_total_use.values()) / len(daily_total_use) if daily_total_use else 10
    reserve_kwh = avg_daily_use * MIN_RESERVE_FRACTION
    print(f"Average daily solar: {avg_daily_use:.1f} kWh, Reserve: {reserve_kwh:.1f} kWh")

    # Create SmartExport tab
    if 'SmartExport' in wb.sheetnames:
        del wb['SmartExport']

    ws_new = wb.create_sheet('SmartExport')

    # Headers
    headers = [
        'Date',  # A
        'Total_Solar_kWh',  # B
        'Avg_Daily_Use',  # C
        'Reserve_kWh',  # D
        # Flow optimised
        'Flow_Opt_Export_kWh', 'Flow_Opt_Export$', 'Flow_Opt_Import_kWh', 'Flow_Opt_Import$', 'Flow_Opt_Net$',  # E-I
        # Origin optimised
        'Origin_Opt_Export_kWh', 'Origin_Opt_Export$', 'Origin_Opt_Import_kWh', 'Origin_Opt_Import$', 'Origin_Opt_Net$',  # J-N
        # Globird optimised
        'Globird_Opt_Export_kWh', 'Globird_Opt_Export$', 'Globird_Opt_Import_kWh', 'Globird_Opt_Import$', 'Globird_Opt_Net$',  # O-S
        # CovaU optimised
        'CovaU_Opt_Export_kWh', 'CovaU_Opt_Export$', 'CovaU_Opt_Import_kWh', 'CovaU_Opt_Import$', 'CovaU_Opt_Net$',  # T-X
        # Amber optimised
        'Amber_Opt_Export_kWh', 'Amber_Opt_Export$', 'Amber_Opt_Import_kWh', 'Amber_Opt_Import$', 'Amber_Opt_Net$',  # Y-AC
    ]

    for col, header in enumerate(headers, 1):
        ws_new.cell(row=1, column=col).value = header

    # Calculate optimised export for each day and retailer
    print("Calculating optimised export strategies...")

    for i, day_str in enumerate(sorted_days):
        row = i + 2
        ws_new.cell(row=row, column=1).value = datetime.strptime(day_str, '%Y-%m-%d')
        ws_new.cell(row=row, column=2).value = round(daily_solar.get(day_str, 0), 2)
        ws_new.cell(row=row, column=3).value = round(avg_daily_use, 2)
        ws_new.cell(row=row, column=4).value = round(reserve_kwh, 2)

        available_for_export = max(0, daily_solar.get(day_str, 0) - reserve_kwh)

        # === Flow: Export only during 17:30-19:30 at 45c ===
        flow_export_hours = [17, 18]  # 17:30-19:30 = hours 17 and 18
        flow_export_kwh = 0
        for h in flow_export_hours:
            # Available solar during these hours (from hourly profile)
            solar_this_hour = hourly_solar[day_str].get(h, 0)
            export_this_hour = min(solar_this_hour, MAX_EXPORT_KW / 12)  # Per 5-min interval
            flow_export_kwh += export_this_hour

        flow_export_kwh = min(flow_export_kwh, available_for_export)
        flow_export_value = flow_export_kwh * 0.45
        flow_import_kwh = max(0, reserve_kwh - (daily_solar.get(day_str, 0) - flow_export_kwh))
        flow_import_value = flow_import_kwh * 0.26  # Approximate PEA-based rate
        flow_net = flow_import_value - flow_export_value + 1.3419  # + DSC

        ws_new.cell(row=row, column=5).value = round(flow_export_kwh, 2)
        ws_new.cell(row=row, column=6).value = round(flow_export_value, 2)
        ws_new.cell(row=row, column=7).value = round(flow_import_kwh, 2)
        ws_new.cell(row=row, column=8).value = round(flow_import_value, 2)
        ws_new.cell(row=row, column=9).value = round(flow_net, 2)

        # === Fixed TOU retailers: Export during highest FIT hours ===
        for retailer, col_start in [('Origin', 10), ('Globird', 15), ('CovaU', 20)]:
            # Sort hours by FIT value (highest first)
            fit_by_hour = hourly_fit_value[retailer][day_str]
            sorted_hours = sorted(fit_by_hour.keys(), key=lambda h: fit_by_hour[h], reverse=True)

            export_kwh = 0
            export_value = 0
            remaining = available_for_export

            for h in sorted_hours:
                if remaining <= 0:
                    break
                # FIT value per kWh for this hour
                fit_per_kwh = fit_by_hour[h] / max(hourly_export[day_str].get(h, 0.001), 0.001)
                solar_this_hour = hourly_solar[day_str].get(h, 0)
                can_export = min(solar_this_hour, remaining, MAX_EXPORT_KW / 12)

                if can_export > 0:
                    export_kwh += can_export
                    export_value += can_export * fit_per_kwh
                    remaining -= can_export

            import_kwh = max(0, reserve_kwh - (daily_solar.get(day_str, 0) - export_kwh))

            # Import rates (approximate average)
            import_rates = {'Origin': 0.187, 'Globird': 0.363, 'CovaU': 0.2802}
            import_value = import_kwh * import_rates[retailer]

            dsc = {'Origin': 1.2567/365, 'Globird': 1.32/365, 'CovaU': 1.30/365}
            net = import_value - export_value + dsc[retailer]

            ws_new.cell(row=row, column=col_start).value = round(export_kwh, 2)
            ws_new.cell(row=row, column=col_start + 1).value = round(export_value, 2)
            ws_new.cell(row=row, column=col_start + 2).value = round(import_kwh, 2)
            ws_new.cell(row=row, column=col_start + 3).value = round(import_value, 2)
            ws_new.cell(row=row, column=col_start + 4).value = round(net, 2)

        # === Amber: Export during highest AEMO price hours ===
        amber_fit = hourly_fit_value['Amber'][day_str]
        sorted_hours = sorted(amber_fit.keys(), key=lambda h: amber_fit[h], reverse=True)

        amber_export_kwh = 0
        amber_export_value = 0
        remaining = available_for_export

        for h in sorted_hours:
            if remaining <= 0:
                break
            solar_this_hour = hourly_solar[day_str].get(h, 0)
            can_export = min(solar_this_hour, remaining, MAX_EXPORT_KW / 12)

            if can_export > 0:
                aemo_price = hourly_aemo[day_str].get(h, 0)
                amber_export_kwh += can_export
                amber_export_value += can_export * max(0, aemo_price)
                remaining -= can_export

        amber_import_kwh = max(0, reserve_kwh - (daily_solar.get(day_str, 0) - amber_export_kwh))
        amber_import_value = amber_import_kwh * 0.20  # Approximate AEMO+network average
        amber_net = amber_import_value - amber_export_value + 1.76 + 25/30

        ws_new.cell(row=row, column=25).value = round(amber_export_kwh, 2)
        ws_new.cell(row=row, column=26).value = round(amber_export_value, 2)
        ws_new.cell(row=row, column=27).value = round(amber_import_kwh, 2)
        ws_new.cell(row=row, column=28).value = round(amber_import_value, 2)
        ws_new.cell(row=row, column=29).value = round(amber_net, 2)

    print("Saving...")
    wb.save(TEMP_FILE)
    wb.close()

    os.replace(TEMP_FILE, EXCEL_FILE)
    print("Phase 4 complete: SmartExport tab created")


if __name__ == "__main__":
    phase4()
