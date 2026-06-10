#!/usr/bin/env python3
"""
Extend Retailer_Rates with Amber entry, Pricing_Model, and Subscription columns.
Phase 1 of multi-retailer comparison.
"""

import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_FILE = "HA Energy 5 Min Pricing.xlsx"

def extend_retailer_rates():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Retailer_Rates']
    
    print("Current Retailer_Rates structure:")
    for row in range(1, ws.max_row + 1):
        vals = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                vals.append(f'{get_column_letter(col)}{row}={val}')
        if vals:
            print(vals)
    
    # Add new headers in columns Q, R, S
    ws['Q1'] = 'Pricing_Model'
    ws['R1'] = 'Subscription_Monthly'
    ws['S1'] = 'Subscription_Daily'
    
    # Add FlowPower entry (new row 5)
    flow_row = 5
    ws[f'A{flow_row}'] = 1  # Id
    ws[f'B{flow_row}'] = 'FlowPower'
    ws[f'C{flow_row}'] = 1.3419  # DSC per day
    ws[f'D{flow_row}'] = 0.34  # Base rate for import calculation
    ws[f'E{flow_row}'] = 0  # Shoulder rate (not used)
    ws[f'F{flow_row}'] = 0  # Peak rate (not used)
    ws[f'G{flow_row}'] = 0.45  # FIT rate during export window
    ws[f'H{flow_row}'] = 0  # FIT Shoulder (not used)
    ws[f'I{flow_row}'] = 0  # FIT Peak (not used)
    ws[f'J{flow_row}'] = 17.5  # Export window start (5:30 PM)
    ws[f'K{flow_row}'] = 19.5  # Export window end (7:30 PM)
    ws[f'L{flow_row}'] = 0  # TOU Peak Start (not used for Flow)
    ws[f'M{flow_row}'] = 0  # TOU Peak End (not used for Flow)
    ws[f'N{flow_row}'] = 0  # FIT Super Peak (not used)
    ws[f'O{flow_row}'] = 0  # FIT Super Peak Start
    ws[f'P{flow_row}'] = 0  # FIT Super Peak End
    ws[f'Q{flow_row}'] = 'hybrid'  # Pricing_Model
    ws[f'R{flow_row}'] = 0  # Subscription_Monthly
    ws[f'S{flow_row}'] = 0  # Subscription_Daily
    
    # Add Amber entry (new row 6)
    amber_row = 6
    ws[f'A{amber_row}'] = 5  # Id
    ws[f'B{amber_row}'] = 'Amber'
    ws[f'C{amber_row}'] = 1.76  # DSC per day
    ws[f'D{amber_row}'] = 0  # Offpeak (variable, not used)
    ws[f'E{amber_row}'] = 0  # Shoulder (variable, not used)
    ws[f'F{amber_row}'] = 0  # Peak (variable, not used)
    ws[f'G{amber_row}'] = 0  # FIT Offpeak (variable, not used)
    ws[f'H{amber_row}'] = 0  # FIT Shoulder (variable, not used)
    ws[f'I{amber_row}'] = 0  # FIT Peak (variable, not used)
    ws[f'J{amber_row}'] = 0  # TOU Offpeak Start (not applicable)
    ws[f'K{amber_row}'] = 0  # TOU Offpeak End (not applicable)
    ws[f'L{amber_row}'] = 0  # TOU Peak Start (not applicable)
    ws[f'M{amber_row}'] = 0  # TOU Peak End (not applicable)
    ws[f'N{amber_row}'] = 0  # FIT Super Peak (not applicable)
    ws[f'O{amber_row}'] = 0  # FIT Super Peak Start (not applicable)
    ws[f'P{amber_row}'] = 0  # FIT Super Peak End (not applicable)
    ws[f'Q{amber_row}'] = 'variable'  # Pricing_Model
    ws[f'R{amber_row}'] = 25  # Subscription_Monthly
    ws[f'S{amber_row}'] = round(25 / 30, 4)  # Subscription_Daily
    
    # Update existing rows with Pricing_Model
    # Row 2 = Origin Loop Max
    ws['Q2'] = 'fixed_tou'
    ws['R2'] = 0
    ws['S2'] = 0
    
    # Row 3 = Globird VPP
    ws['Q3'] = 'fixed_tou'
    ws['R3'] = 0
    ws['S3'] = 0
    
    # Row 4 = CovaU SolarMax
    ws['Q4'] = 'fixed_tou'
    ws['R4'] = 0
    ws['S4'] = 0
    
    print("\nUpdated Retailer_Rates:")
    for row in range(1, ws.max_row + 1):
        vals = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                vals.append(f'{get_column_letter(col)}{row}={val}')
        if vals:
            print(vals)
    
    wb.save(EXCEL_FILE)
    print("\nSaved successfully!")

if __name__ == "__main__":
    extend_retailer_rates()
