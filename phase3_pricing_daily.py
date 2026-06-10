#!/usr/bin/env python3
"""
Phase 3: Create Retailer_Comparison tab with per-retailer daily data.
Keeps existing PricingDaily intact. New tab provides unified comparison.
"""
import openpyxl
import os
import shutil
from datetime import datetime, date

EXCEL_FILE = "HA Energy 5 Min Pricing.xlsx"
TEMP_FILE = "HA Energy 5 Min Pricing temp.xlsx"

RETAILERS = {
    'Flow': {'dsc_daily': 1.3419, 'sub_daily': 0},
    'Origin': {'dsc_daily': round(1.2567 / 365, 6), 'sub_daily': 0},
    'Globird': {'dsc_daily': round(1.32 / 365, 6), 'sub_daily': 0},
    'CovaU': {'dsc_daily': round(1.30 / 365, 6), 'sub_daily': 0},
    'Amber': {'dsc_daily': 1.76, 'sub_daily': round(25 / 30, 4)},
}


def phase3():
    print("Loading workbook...")
    shutil.copy2(EXCEL_FILE, TEMP_FILE)
    wb = openpyxl.load_workbook(TEMP_FILE, data_only=True)

    ws_enh = wb['Pricing5MinEnhanced']
    enh_max = ws_enh.max_row
    print(f"Pricing5MinEnhanced: {enh_max} rows")

    # Build daily aggregation
    daily_data = {}
    batch_size = 10000
    print("Building daily aggregation...")
    for start in range(2, enh_max + 1, batch_size):
        end = min(start + batch_size - 1, enh_max)
        if start % 30000 == 2:
            print(f"  Rows {start}-{end}...")

        for row in range(start, end + 1):
            date_val = ws_enh.cell(row=row, column=17).value  # Q = Date
            if date_val is None:
                continue

            if hasattr(date_val, 'date'):
                day_str = str(date_val.date())
            elif isinstance(date_val, datetime):
                day_str = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, date):
                day_str = date_val.strftime('%Y-%m-%d')
            else:
                day_str = str(date_val)[:10]

            if day_str not in daily_data:
                daily_data[day_str] = {
                    'origin_import': 0, 'origin_export': 0,
                    'globird_import': 0, 'globird_export': 0,
                    'covau_import': 0, 'covau_export': 0,
                    'amber_import': 0, 'amber_export': 0,
                }

            d = daily_data[day_str]

            # Origin Import$ (col AD=30) and Export$ (col AE=31)
            val = ws_enh.cell(row=row, column=30).value
            if isinstance(val, (int, float)):
                d['origin_import'] += val
            val = ws_enh.cell(row=row, column=31).value
            if isinstance(val, (int, float)):
                d['origin_export'] += val

            # Globird Import$ (col AG=33) and Export$ (col AH=34)
            val = ws_enh.cell(row=row, column=33).value
            if isinstance(val, (int, float)):
                d['globird_import'] += val
            val = ws_enh.cell(row=row, column=34).value
            if isinstance(val, (int, float)):
                d['globird_export'] += val

            # CovaU Import$ (col AJ=36) and Export$ (col AK=37)
            val = ws_enh.cell(row=row, column=36).value
            if isinstance(val, (int, float)):
                d['covau_import'] += val
            val = ws_enh.cell(row=row, column=37).value
            if isinstance(val, (int, float)):
                d['covau_export'] += val

            # Amber Import$ = Incl NW Price * imported (col T=20)
            val = ws_enh.cell(row=row, column=20).value
            if isinstance(val, (int, float)):
                d['amber_import'] += val

            # Amber Export$ (col AL=38)
            val = ws_enh.cell(row=row, column=38).value
            if isinstance(val, (int, float)):
                d['amber_export'] += val

    print(f"Aggregated {len(daily_data)} days")

    # Also read Flow daily data from PricingDaily
    ws_daily = wb['PricingDaily']
    flow_daily = {}
    for row in range(2, ws_daily.max_row + 1):
        date_val = ws_daily.cell(row=row, column=1).value
        if date_val is None:
            continue
        if hasattr(date_val, 'date'):
            day_str = str(date_val.date())
        elif isinstance(date_val, datetime):
            day_str = date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, date):
            day_str = date_val.strftime('%Y-%m-%d')
        else:
            day_str = str(date_val)[:10]

        flow_import = ws_daily.cell(row=row, column=2).value or 0
        flow_export = ws_daily.cell(row=row, column=3).value or 0

        flow_daily[day_str] = {
            'import': float(flow_import) if isinstance(flow_import, (int, float)) else 0,
            'export': float(flow_export) if isinstance(flow_export, (int, float)) else 0,
        }

    print(f"Read {len(flow_daily)} days from PricingDaily")

    # Create new Retailer_Comparison tab
    if 'Retailer_Comparison' in wb.sheetnames:
        del wb['Retailer_Comparison']

    ws_new = wb.create_sheet('Retailer_Comparison')

    # Headers
    headers = [
        'Date',  # A
        'Flow_Import$', 'Flow_Export$', 'Flow_DSC', 'Flow_Net$',  # B-E
        'Origin_Import$', 'Origin_Export$', 'Origin_DSC', 'Origin_Net$',  # F-I
        'Globird_Import$', 'Globird_Export$', 'Globird_DSC', 'Globird_Net$',  # J-M
        'CovaU_Import$', 'CovaU_Export$', 'CovaU_DSC', 'CovaU_Net$',  # N-Q
        'Amber_Import$', 'Amber_Export$', 'Amber_DSC', 'Amber_Sub', 'Amber_Net$',  # R-V
        'Cheapest_Retailer', 'Savings_vs_Flow'  # W-X
    ]

    for col, header in enumerate(headers, 1):
        ws_new.cell(row=1, column=col).value = header

    # Write data rows
    sorted_days = sorted(daily_data.keys(), reverse=True)
    print(f"Writing {len(sorted_days)} rows to Retailer_Comparison...")

    for i, day_str in enumerate(sorted_days):
        row = i + 2
        d = daily_data[day_str]

        # Parse date
        try:
            dt = datetime.strptime(day_str, '%Y-%m-%d')
            ws_new.cell(row=row, column=1).value = dt
        except ValueError:
            ws_new.cell(row=row, column=1).value = day_str

        # Flow (from PricingDaily)
        fd = flow_daily.get(day_str, {'import': 0, 'export': 0})
        flow_import = fd['import']
        flow_export = abs(fd['export'])  # Make positive (export credit)
        flow_dsc = RETAILERS['Flow']['dsc_daily']
        flow_net = flow_import - flow_export + flow_dsc

        ws_new.cell(row=row, column=2).value = round(flow_import, 4)
        ws_new.cell(row=row, column=3).value = round(flow_export, 4)
        ws_new.cell(row=row, column=4).value = round(flow_dsc, 4)
        ws_new.cell(row=row, column=5).value = round(flow_net, 4)

        # Origin
        origin_import = d['origin_import']
        origin_export = d['origin_export']
        origin_dsc = RETAILERS['Origin']['dsc_daily']
        origin_net = origin_import - origin_export + origin_dsc

        ws_new.cell(row=row, column=6).value = round(origin_import, 4)
        ws_new.cell(row=row, column=7).value = round(origin_export, 4)
        ws_new.cell(row=row, column=8).value = round(origin_dsc, 4)
        ws_new.cell(row=row, column=9).value = round(origin_net, 4)

        # Globird
        globird_import = d['globird_import']
        globird_export = d['globird_export']
        globird_dsc = RETAILERS['Globird']['dsc_daily']
        globird_net = globird_import - globird_export + globird_dsc

        ws_new.cell(row=row, column=10).value = round(globird_import, 4)
        ws_new.cell(row=row, column=11).value = round(globird_export, 4)
        ws_new.cell(row=row, column=12).value = round(globird_dsc, 4)
        ws_new.cell(row=row, column=13).value = round(globird_net, 4)

        # CovaU
        covau_import = d['covau_import']
        covau_export = d['covau_export']
        covau_dsc = RETAILERS['CovaU']['dsc_daily']
        covau_net = covau_import - covau_export + covau_dsc

        ws_new.cell(row=row, column=14).value = round(covau_import, 4)
        ws_new.cell(row=row, column=15).value = round(covau_export, 4)
        ws_new.cell(row=row, column=16).value = round(covau_dsc, 4)
        ws_new.cell(row=row, column=17).value = round(covau_net, 4)

        # Amber
        amber_import = d['amber_import']
        amber_export = d['amber_export']
        amber_dsc = RETAILERS['Amber']['dsc_daily']
        amber_sub = RETAILERS['Amber']['sub_daily']
        amber_net = amber_import - amber_export + amber_dsc + amber_sub

        ws_new.cell(row=row, column=18).value = round(amber_import, 4)
        ws_new.cell(row=row, column=19).value = round(amber_export, 4)
        ws_new.cell(row=row, column=20).value = round(amber_dsc, 4)
        ws_new.cell(row=row, column=21).value = round(amber_sub, 4)
        ws_new.cell(row=row, column=22).value = round(amber_net, 4)

        # Find cheapest retailer
        costs = {
            'Flow': flow_net,
            'Origin': origin_net,
            'Globird': globird_net,
            'CovaU': covau_net,
            'Amber': amber_net,
        }
        cheapest = min(costs, key=costs.get)
        savings = flow_net - costs[cheapest]

        ws_new.cell(row=row, column=23).value = cheapest
        ws_new.cell(row=row, column=24).value = round(savings, 4)

    # Add summary section at bottom
    summary_row = len(sorted_days) + 3
    ws_new.cell(row=summary_row, column=1).value = 'SUMMARY'
    ws_new.cell(row=summary_row + 1, column=1).value = 'Total Days'

    # Calculate totals
    totals = {r: {'import': 0, 'export': 0, 'net': 0} for r in RETAILERS}
    for day_str in sorted_days:
        d = daily_data[day_str]
        fd = flow_daily.get(day_str, {'import': 0, 'export': 0})

        totals['Flow']['import'] += fd['import']
        totals['Flow']['export'] += abs(fd['export'])
        totals['Flow']['net'] += fd['import'] - abs(fd['export']) + RETAILERS['Flow']['dsc_daily']

        totals['Origin']['import'] += d['origin_import']
        totals['Origin']['export'] += d['origin_export']
        totals['Origin']['net'] += d['origin_import'] - d['origin_export'] + RETAILERS['Origin']['dsc_daily']

        totals['Globird']['import'] += d['globird_import']
        totals['Globird']['export'] += d['globird_export']
        totals['Globird']['net'] += d['globird_import'] - d['globird_export'] + RETAILERS['Globird']['dsc_daily']

        totals['CovaU']['import'] += d['covau_import']
        totals['CovaU']['export'] += d['covau_export']
        totals['CovaU']['net'] += d['covau_import'] - d['covau_export'] + RETAILERS['CovaU']['dsc_daily']

        totals['Amber']['import'] += d['amber_import']
        totals['Amber']['export'] += d['amber_export']
        totals['Amber']['net'] += d['amber_import'] - d['amber_export'] + RETAILERS['Amber']['dsc_daily'] + RETAILERS['Amber']['sub_daily']

    r = summary_row + 2
    ws_new.cell(row=r, column=1).value = 'Retailer'
    ws_new.cell(row=r, column=2).value = 'Total Import'
    ws_new.cell(row=r, column=3).value = 'Total Export'
    ws_new.cell(row=r, column=4).value = 'Total Net Cost'
    ws_new.cell(row=r, column=5).value = 'Avg Daily Net'

    for retailer_name in ['Flow', 'Origin', 'Globird', 'CovaU', 'Amber']:
        r += 1
        t = totals[retailer_name]
        ws_new.cell(row=r, column=1).value = retailer_name
        ws_new.cell(row=r, column=2).value = round(t['import'], 2)
        ws_new.cell(row=r, column=3).value = round(t['export'], 2)
        ws_new.cell(row=r, column=4).value = round(t['net'], 2)
        ws_new.cell(row=r, column=5).value = round(t['net'] / len(sorted_days), 2) if sorted_days else 0

    print("Saving...")
    wb.save(TEMP_FILE)
    wb.close()

    os.replace(TEMP_FILE, EXCEL_FILE)
    print("Phase 3 complete: Retailer_Comparison tab created")


if __name__ == "__main__":
    phase3()
